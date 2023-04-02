import openpyxl
wb = openpyxl.load_workbook('C:/Users/91761/Documents/produceSales.xlsx')
sheet = wb['Sheet1']
PRICE_UPDATES = {'Garlic': 2.00,
                 'Celery': 4.56,
                'Lemon': 5.90}
for rowNum in range(2,10):
	produceName = sheet.cell(row=rowNum, column=1).value
	if produceName in PRICE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updatedProduceSales.xlsx')


